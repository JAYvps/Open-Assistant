#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
markdown表格转Excel工具
功能：将markdown文件中的表格内容提取并转换为Excel文件，保存到桌面并自动打开
"""

import os
import sys
import subprocess
import markdown
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment

def markdown_to_excel_main(input_data):
    """
    将markdown内容（文件或字符串）中的表格转换为Excel文件，保存到桌面并自动打开
    
    Args:
        input_data (str): markdown文件路径或markdown字符串内容
    
    Returns:
        str or None: 生成的Excel文件路径，如果处理失败则返回None
    """
    # 获取桌面路径
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    
    # 确定输出文件名，避免覆盖
    base_filename = "new.xlsx"
    output_path = os.path.join(desktop_path, base_filename)
    counter = 1
    while os.path.exists(output_path):
        output_path = os.path.join(desktop_path, f"new_{counter}.xlsx")
        counter += 1
    
    # 判断输入是文件路径还是markdown字符串
    if os.path.exists(input_data):
        # 输入是文件路径，读取文件内容
        try:
            with open(input_data, 'r', encoding='utf-8') as f:
                md_content = f.read()
            print(f"成功读取文件：{input_data}")
        except Exception as e:
            print(f"读取markdown文件时出错：{str(e)}")
            return None
    else:
        # 输入是markdown字符串
        md_content = input_data
        print("正在处理markdown字符串内容...")
    
    # 使用markdown库解析表格，启用表格扩展
    html_content = markdown.markdown(md_content, extensions=['tables'])
    soup = BeautifulSoup(html_content, 'html.parser')
    tables = soup.find_all('table')
    
    if not tables:
        print("未在markdown文件中找到表格")
        return "未在markdown文件中找到表格"
    
    # 创建Excel工作簿
    workbook = openpyxl.Workbook()
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 处理每个表格
    for table_idx, table in enumerate(tables):
        if table_idx == 0:
            # 第一个表格，使用默认的第一个工作表
            worksheet = workbook.active
            worksheet.title = "表格1"
        else:
            # 非第一个表格，创建新的工作表
            worksheet = workbook.create_sheet(title=f"表格{table_idx+1}")
        
        # 获取所有行
        rows = table.find_all('tr')
        
        for row_idx, row in enumerate(rows):
            # 获取所有单元格
            cells = row.find_all(['th', 'td'])
            
            for col_idx, cell in enumerate(cells):
                # 获取单元格内容
                cell_text = cell.get_text().strip()
                
                # 写入Excel单元格
                excel_cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_text)
                
                # 设置单元格样式
                if cell.name == 'th':  # 表头
                    excel_cell.font = header_font
                    excel_cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    excel_cell.alignment = header_alignment
                else:  # 数据单元格
                    # 尝试将数字字符串转换为数字
                    try:
                        # 尝试转换为浮点数
                        if '.' in cell_text:
                            excel_cell.value = float(cell_text)
                        else:
                            # 尝试转换为整数
                            excel_cell.value = int(cell_text)
                    except ValueError:
                        # 不是数字，保持为文本
                        pass
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 保存Excel文件
    try:
        workbook.save(output_path)
        print(f"成功生成Excel文件：{output_path}")
        
        # 打开生成的Excel文件
        try:
            if os.name == 'nt':  # Windows
                os.startfile(output_path)
            elif os.name == 'posix':  # macOS or Linux
                subprocess.call(['open', output_path]) if sys.platform == 'darwin' else subprocess.call(['xdg-open', output_path])
            print("已自动打开Excel文件")
        except Exception as e:
            print(f"尝试打开Excel文件时出错：{str(e)}")
            print(f"请手动打开文件：{output_path}")
        return output_path
    except Exception as e:
        print(f"保存Excel文件时出错：{str(e)}")
        return None


if __name__ == "__main__":
    # 获取命令行参数，如果没有提供则使用默认测试文件
    if len(sys.argv) > 1:
        input_data = sys.argv[1]
    else:
        # 默认使用测试文件
        input_data = "test_markdown_table.md"
    
    # 调用主函数并获取返回值
    excel_file_path = markdown_to_excel_main(input_data)
    
    # 可以使用返回的文件路径进行后续操作
    if excel_file_path:
        print(f"\n函数返回的Excel文件路径：{excel_file_path}")
        print(f"文件是否存在：{os.path.exists(excel_file_path)}")
    else:
        print("\n函数执行失败，未生成Excel文件")
    
    # 示例：如何直接传入markdown字符串
    # markdown_str = "| 姓名 | 年龄 | 职业 |\n|------|------|------|\n| 张三 | 25 | 工程师 |\n| 李四 | 30 | 设计师 |"
    # excel_path = markdown_to_excel_main(markdown_str)
    # print(f"从字符串生成的Excel路径：{excel_path}")